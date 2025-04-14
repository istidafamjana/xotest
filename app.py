import os
import json
import hashlib
import uuid
import time
import re
import mimetypes
from datetime import datetime, timedelta
from threading import Lock, Thread
from functools import wraps
import tempfile
import urllib.request
import logging
from flask import Flask, request, jsonify, render_template_string, redirect, url_for, session, flash, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import requests
import google.generativeai as genai
from PIL import Image
import pytesseract
from pdf2image import convert_from_bytes
import io

# تكوين التطبيق الأساسي
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-123')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=5)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# تكوين السجلات
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# التوكنات والمفاتيح
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', 'AIzaSyA1TKhF1NQskLCqXR3O_cpISpTn9I8R-IU')

# تهيئة نموذج Gemini
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

# تخزين البيانات
DATA_FILE = "users_data.json"
conversations = {}
users = {}
CONVERSATION_TIMEOUT = 5 * 60 * 60  # 5 ساعات بالثواني
data_lock = Lock()

# إنشاء مجلد التحميلات إذا لم يكن موجوداً
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# تعريف القوالب المضمنة
TEMPLATES = {
    'base.html': '''
<!DOCTYPE html>
<html dir="rtl" lang="ar" data-bs-theme="{{ theme }}">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ title }} - OTH AI</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/styles/atom-one-dark.min.css" rel="stylesheet">
    <style>
        :root {
            --primary-color: #6e8efb;
            --secondary-color: #a777e3;
            --dark-bg: #1a1a2e;
            --dark-card: #16213e;
            --dark-text: #e6e6e6;
        }
        
        body {
            background-color: {% if theme == 'dark' %}var(--dark-bg){% else %}#f8f9fa{% endif %};
            color: {% if theme == 'dark' %}var(--dark-text){% endif %};
            transition: all 0.3s ease;
        }
        
        .hero-section {
            background: linear-gradient(135deg, var(--primary-color), var(--secondary-color));
        }
        
        .feature-icon {
            font-size: 2rem;
            color: var(--primary-color);
        }
        
        .card {
            background-color: {% if theme == 'dark' %}var(--dark-card){% else %}#ffffff{% endif %};
            border: none;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            transition: transform 0.3s ease;
        }
        
        .card:hover {
            transform: translateY(-5px);
        }
        
        .chat-container {
            max-width: 900px;
            margin: 0 auto;
            border-radius: 10px;
        }
        
        .chat-messages {
            height: 60vh;
            overflow-y: auto;
            scroll-behavior: smooth;
        }
        
        .message {
            max-width: 80%;
            margin-bottom: 15px;
            padding: 12px 15px;
            border-radius: 12px;
            animation: fadeIn 0.3s ease;
        }
        
        .user-message {
            margin-left: auto;
            background-color: {% if theme == 'dark' %}#2a3f5f{% else %}#e3f2fd{% endif %};
        }
        
        .bot-message {
            margin-right: auto;
            background-color: {% if theme == 'dark' %}#3a3a4e{% else %}#f1f1f1{% endif %};
        }
        
        .code-block {
            position: relative;
            margin: 10px 0;
            border-radius: 8px;
            overflow: hidden;
        }
        
        .code-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 5px 10px;
            background-color: {% if theme == 'dark' %}#2d2d2d{% else %}#f5f5f5{% endif %};
            font-family: monospace;
        }
        
        .copy-btn {
            background: none;
            border: none;
            color: {% if theme == 'dark' %}#ffffff{% else %}#333333{% endif %};
            cursor: pointer;
        }
        
        pre {
            margin: 0;
            padding: 10px;
            overflow-x: auto;
        }
        
        .typing-effect {
            display: inline-block;
            overflow: hidden;
            white-space: nowrap;
            animation: typing 1s steps(40, end);
        }
        
        .file-upload-wrapper {
            position: relative;
            margin-bottom: 15px;
        }
        
        .file-upload-label {
            display: block;
            padding: 15px;
            border: 2px dashed #ccc;
            border-radius: 8px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        
        .file-upload-label:hover {
            border-color: var(--primary-color);
        }
        
        .file-upload-input {
            position: absolute;
            left: 0;
            top: 0;
            opacity: 0;
            width: 100%;
            height: 100%;
            cursor: pointer;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes typing {
            from { width: 0 }
            to { width: 100% }
        }
        
        .language-badge {
            font-size: 0.8rem;
            padding: 3px 8px;
            border-radius: 4px;
            background-color: var(--primary-color);
            color: white;
        }
    </style>
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-dark">
        <div class="container">
            <a class="navbar-brand d-flex align-items-center" href="/">
                <i class="bi bi-robot me-2"></i>
                OTH AI
            </a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav me-auto">
                    <li class="nav-item"><a class="nav-link" href="/chat"><i class="bi bi-chat-left-text me-1"></i> المحادثة</a></li>
                    <li class="nav-item"><a class="nav-link" href="/features"><i class="bi bi-stars me-1"></i> الميزات</a></li>
                    <li class="nav-item"><a class="nav-link" href="/about"><i class="bi bi-info-circle me-1"></i> حول</a></li>
                </ul>
                <div class="d-flex align-items-center">
                    <button id="theme-toggle" class="btn btn-sm btn-outline-light me-2">
                        <i class="bi {% if theme == 'dark' %}bi-sun{% else %}bi-moon{% endif %}"></i>
                    </button>
                    {% if 'user_id' in session %}
                        <div class="dropdown">
                            <button class="btn btn-outline-light dropdown-toggle" type="button" id="userDropdown" data-bs-toggle="dropdown">
                                <i class="bi bi-person-circle me-1"></i> {{ session['username'] }}
                            </button>
                            <ul class="dropdown-menu dropdown-menu-end">
                                <li><a class="dropdown-item" href="/profile"><i class="bi bi-person me-2"></i> الملف الشخصي</a></li>
                                <li><a class="dropdown-item" href="/settings"><i class="bi bi-gear me-2"></i> الإعدادات</a></li>
                                <li><hr class="dropdown-divider"></li>
                                <li><a class="dropdown-item" href="/logout"><i class="bi bi-box-arrow-right me-2"></i> تسجيل الخروج</a></li>
                            </ul>
                        </div>
                    {% else %}
                        <a href="/login" class="btn btn-outline-light me-2">تسجيل الدخول</a>
                        <a href="/register" class="btn btn-primary">إنشاء حساب</a>
                    {% endif %}
                </div>
            </div>
        </div>
    </nav>

    <div class="container my-4">
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
                        {{ message }}
                        <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        
        {% block content %}{% endblock %}
    </div>

    <footer class="bg-dark text-white py-4 mt-5">
        <div class="container">
            <div class="row">
                <div class="col-md-4">
                    <h5><i class="bi bi-robot"></i> OTH AI</h5>
                    <p>منصة الذكاء الاصطناعي المتكاملة تعمل بنموذج Gemini 1.5 Flash لتقديم أفضل تجربة محادثة.</p>
                </div>
                <div class="col-md-4">
                    <h5>روابط سريعة</h5>
                    <ul class="list-unstyled">
                        <li><a href="/features" class="text-white">الميزات</a></li>
                        <li><a href="/about" class="text-white">حول</a></li>
                        <li><a href="/privacy" class="text-white">الخصوصية</a></li>
                    </ul>
                </div>
                <div class="col-md-4">
                    <h5>اتصل بنا</h5>
                    <ul class="list-unstyled">
                        <li><i class="bi bi-envelope me-2"></i> contact@othai.com</li>
                        <li><i class="bi bi-twitter me-2"></i> @othai_support</li>
                    </ul>
                </div>
            </div>
            <hr class="my-4">
            <div class="text-center">
                <p class="mb-0">© 2023 OTH AI. جميع الحقوق محفوظة.</p>
            </div>
        </div>
    </footer>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/highlight.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/languages/python.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/languages/javascript.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/languages/htmlbars.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/languages/css.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/languages/json.min.js"></script>
    <script>
        // تنشيط تظليل الكود
        document.addEventListener('DOMContentLoaded', function() {
            document.querySelectorAll('pre code').forEach((block) => {
                hljs.highlightElement(block);
            });
        });
        
        // نسخ الكود
        function copyCode(button) {
            const codeBlock = button.parentElement.nextElementSibling;
            const codeText = codeBlock.innerText;
            navigator.clipboard.writeText(codeText).then(() => {
                button.innerHTML = '<i class="bi bi-check"></i> نسخ!';
                setTimeout(() => {
                    button.innerHTML = '<i class="bi bi-clipboard"></i> نسخ';
                }, 2000);
            });
        }
        
        // تبديل السمة
        const themeToggle = document.getElementById('theme-toggle');
        if (themeToggle) {
            themeToggle.addEventListener('click', function() {
                const html = document.documentElement;
                const currentTheme = html.getAttribute('data-bs-theme');
                const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
                
                html.setAttribute('data-bs-theme', newTheme);
                localStorage.setItem('theme', newTheme);
                
                this.innerHTML = `<i class="bi ${newTheme === 'dark' ? 'bi-sun' : 'bi-moon'}"></i>`;
                
                // تحديث أيقونة الزر
                const icon = this.querySelector('i');
                icon.className = `bi ${newTheme === 'dark' ? 'bi-sun' : 'bi-moon'}`;
            });
        }
        
        // استعادة السمة من localStorage
        const savedTheme = localStorage.getItem('theme') || 'light';
        document.documentElement.setAttribute('data-bs-theme', savedTheme);
    </script>
    
    {% block scripts %}{% endblock %}
</body>
</html>
    ''',
    
    'index.html': '''
{% extends "base.html" %}

{% block content %}
<section class="hero-section text-white py-5">
    <div class="container py-5 text-center">
        <h1 class="display-4 fw-bold">منصة الذكاء الاصطناعي OTH</h1>
        <p class="lead">تجربة محادثة متقدمة مع Gemini 1.5 Flash</p>
        {% if 'user_id' not in session %}
            <a href="/register" class="btn btn-light btn-lg mt-3">ابدأ الآن</a>
        {% else %}
            <a href="/chat" class="btn btn-light btn-lg mt-3">اذهب إلى المحادثة</a>
        {% endif %}
    </div>
</section>

<div class="container py-5">
    <div class="row g-4">
        <div class="col-md-4">
            <div class="card h-100">
                <div class="card-body text-center">
                    <div class="feature-icon mb-3"><i class="bi bi-robot"></i></div>
                    <h3>ذكاء اصطناعي متقدم</h3>
                    <p>محادثات ذكية مع نموذج Gemini 1.5 Flash من جوجل</p>
                </div>
            </div>
        </div>
        <div class="col-md-4">
            <div class="card h-100">
                <div class="card-body text-center">
                    <div class="feature-icon mb-3"><i class="bi bi-file-earmark-text"></i></div>
                    <h3>تحليل الملفات</h3>
                    <p>يدعم PDF، Word، Excel، PowerPoint والصور</p>
                </div>
            </div>
        </div>
        <div class="col-md-4">
            <div class="card h-100">
                <div class="card-body text-center">
                    <div class="feature-icon mb-3"><i class="bi bi-code-square"></i></div>
                    <h3>تحليل الكود</h3>
                    <p>فهم وتفسير الكود البرمجي بمختلف اللغات</p>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    ''',
    
    'login.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5">
    <div class="login-card card shadow">
        <div class="card-body p-5">
            <h2 class="card-title text-center mb-4">تسجيل الدخول</h2>
            
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
                            {{ message }}
                            <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
            
            <form method="POST" action="/login">
                <input type="hidden" name="next" value="{{ request.args.get('next', '') }}">
                
                <div class="mb-3">
                    <label for="username" class="form-label">اسم المستخدم</label>
                    <input type="text" class="form-control" id="username" name="username" required>
                </div>
                
                <div class="mb-4">
                    <label for="password" class="form-label">كلمة المرور</label>
                    <input type="password" class="form-control" id="password" name="password" required>
                </div>
                
                <button type="submit" class="btn btn-primary w-100 py-2">تسجيل الدخول</button>
            </form>
            
            <div class="text-center mt-3">
                <p>ليس لديك حساب؟ <a href="/register">إنشاء حساب جديد</a></p>
                <p><a href="/">العودة للصفحة الرئيسية</a></p>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    ''',
    
    'register.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5">
    <div class="register-card card shadow">
        <div class="card-body p-5">
            <h2 class="card-title text-center mb-4">إنشاء حساب جديد</h2>
            
            {% with messages = get_flashed_messages(with_categories=true) %}
                {% if messages %}
                    {% for category, message in messages %}
                        <div class="alert alert-{{ category }} alert-dismissible fade show" role="alert">
                            {{ message }}
                            <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>
                        </div>
                    {% endfor %}
                {% endif %}
            {% endwith %}
            
            <form method="POST" action="/register">
                <div class="mb-3">
                    <label for="username" class="form-label">اسم المستخدم</label>
                    <input type="text" class="form-control" id="username" name="username" required>
                    <div class="form-text">يجب أن يكون 4 أحرف على الأقل</div>
                </div>
                
                <div class="mb-4">
                    <label for="password" class="form-label">كلمة المرور</label>
                    <input type="password" class="form-control" id="password" name="password" required>
                    <div class="form-text">يجب أن تكون 6 أحرف على الأقل</div>
                </div>
                
                <button type="submit" class="btn btn-primary w-100 py-2">إنشاء حساب</button>
            </form>
            
            <div class="text-center mt-3">
                <p>لديك حساب بالفعل؟ <a href="/login">تسجيل الدخول</a></p>
                <p><a href="/">العودة للصفحة الرئيسية</a></p>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    ''',
    
    'chat.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-4">
    <div class="chat-container card shadow">
        <div class="card-header bg-primary text-white d-flex justify-content-between align-items-center">
            <h5 class="mb-0"><i class="bi bi-chat-left-text me-2"></i> محادثة مع OTH AI</h5>
            <div>
                <button id="new-chat-btn" class="btn btn-sm btn-light">
                    <i class="bi bi-plus-lg"></i> محادثة جديدة
                </button>
            </div>
        </div>
        
        <div class="card-body">
            <div id="chat-messages" class="chat-messages mb-3 p-3">
                <!-- سيتم ملء الرسائل هنا عبر JavaScript -->
            </div>
            
            <div class="file-upload-wrapper mb-3">
                <label for="file-upload" class="file-upload-label">
                    <i class="bi bi-upload fs-4"></i>
                    <div>اسحب وأسقط الملفات هنا أو انقر للاختيار</div>
                    <small class="text-muted">يدعم: الصور، PDF، Word، Excel (حتى 16MB)</small>
                </label>
                <input type="file" id="file-upload" class="file-upload-input" accept="image/*,.pdf,.doc,.docx,.xls,.xlsx">
            </div>
            
            <form id="chat-form" class="d-flex">
                <input type="text" id="user-input" class="form-control me-2" placeholder="اكتب رسالتك هنا..." required>
                <button type="submit" class="btn btn-primary">
                    <i class="bi bi-send"></i> إرسال
                </button>
            </form>
        </div>
    </div>
</div>
{% endblock %}

{% block scripts %}
<script>
    document.addEventListener('DOMContentLoaded', function() {
        const chatForm = document.getElementById('chat-form');
        const userInput = document.getElementById('user-input');
        const chatMessages = document.getElementById('chat-messages');
        const fileUpload = document.getElementById('file-upload');
        const newChatBtn = document.getElementById('new-chat-btn');
        
        // تحميل المحادثة السابقة إذا وجدت
        loadPreviousMessages();
        
        function addMessage(sender, message, isCode = false, language = null) {
            const messageDiv = document.createElement('div');
            messageDiv.className = `message ${sender}-message`;
            
            if (isCode) {
                const codeContainer = document.createElement('div');
                codeContainer.className = 'code-block';
                
                const codeHeader = document.createElement('div');
                codeHeader.className = 'code-header';
                
                if (language) {
                    const langBadge = document.createElement('span');
                    langBadge.className = 'language-badge';
                    langBadge.textContent = language;
                    codeHeader.appendChild(langBadge);
                }
                
                const copyButton = document.createElement('button');
                copyButton.className = 'copy-btn ms-auto';
                copyButton.innerHTML = '<i class="bi bi-clipboard"></i> نسخ';
                copyButton.onclick = function() { copyCode(this); };
                codeHeader.appendChild(copyButton);
                
                const codeElement = document.createElement('pre');
                const codeBlock = document.createElement('code');
                codeBlock.className = language ? `language-${language}` : '';
                codeBlock.textContent = message;
                codeElement.appendChild(codeBlock);
                
                codeContainer.appendChild(codeHeader);
                codeContainer.appendChild(codeElement);
                messageDiv.appendChild(codeContainer);
            } else {
                messageDiv.textContent = message;
            }
            
            chatMessages.appendChild(messageDiv);
            chatMessages.scrollTop = chatMessages.scrollHeight;
        }
        
        function addTypingMessage() {
            const typingDiv = document.createElement('div');
            typingDiv.className = 'message bot-message';
            typingDiv.id = 'typing-message';
            
            const typingText = document.createElement('span');
            typingText.className = 'typing-effect';
            typingText.textContent = 'يكتب...';
            
            typingDiv.appendChild(typingText);
            chatMessages.appendChild(typingDiv);
            chatMessages.scrollTop = chatMessages.scrollHeight;
            
            return typingDiv;
        }
        
        function removeTypingMessage() {
            const typingDiv = document.getElementById('typing-message');
            if (typingDiv) {
                typingDiv.remove();
            }
        }
        
        function detectCode(message) {
            // اكتشاف الأكواد في الرسالة
            const codeBlocks = [];
            const codeRegex = /```(\w*)\n([\s\S]*?)\n```/g;
            let match;
            let lastIndex = 0;
            let processedMessage = '';
            
            while ((match = codeRegex.exec(message)) !== null) {
                const language = match[1] || 'plaintext';
                const code = match[2];
                
                // النص قبل الكود
                if (match.index > lastIndex) {
                    processedMessage += message.substring(lastIndex, match.index);
                }
                
                codeBlocks.push({
                    language: language,
                    code: code,
                    position: processedMessage.length
                });
                
                // نضع علامة مكان الكود
                processedMessage += `\x1Bcode${codeBlocks.length - 1}\x1B`;
                lastIndex = codeRegex.lastIndex;
            }
            
            // النص المتبقي بعد آخر كود
            if (lastIndex < message.length) {
                processedMessage += message.substring(lastIndex);
            }
            
            return { text: processedMessage, codeBlocks: codeBlocks };
        }
        
        function processMessageWithCode(message) {
            const { text, codeBlocks } = detectCode(message);
            
            if (codeBlocks.length === 0) {
                return [{ type: 'text', content: message }];
            }
            
            const parts = [];
            let lastPos = 0;
            
            // تقسيم الرسالة إلى أجزاء نصية وكود
            for (let i = 0; i < codeBlocks.length; i++) {
                const codePos = text.indexOf(`\x1Bcode${i}\x1B`);
                
                // النص قبل الكود
                if (codePos > lastPos) {
                    parts.push({
                        type: 'text',
                        content: text.substring(lastPos, codePos)
                    });
                }
                
                // إضافة الكود
                parts.push({
                    type: 'code',
                    content: codeBlocks[i].code,
                    language: codeBlocks[i].language
                });
                
                lastPos = codePos + `\x1Bcode${i}\x1B`.length;
            }
            
            // النص بعد آخر كود
            if (lastPos < text.length) {
                parts.push({
                    type: 'text',
                    content: text.substring(lastPos)
                });
            }
            
            return parts;
        }
        
        function loadPreviousMessages() {
            fetch('/api/conversation')
                .then(response => response.json())
                .then(data => {
                    if (data.messages) {
                        data.messages.forEach(msg => {
                            if (msg.sender === 'user') {
                                addMessage('user', msg.content);
                            } else {
                                const messageParts = processMessageWithCode(msg.content);
                                messageParts.forEach(part => {
                                    if (part.type === 'text') {
                                        addMessage('bot', part.content);
                                    } else if (part.type === 'code') {
                                        addMessage('bot', part.content, true, part.language);
                                    }
                                });
                            }
                        });
                    }
                })
                .catch(error => {
                    console.error('Error loading conversation:', error);
                });
        }
        
        function sendMessageToServer(message, file = null) {
            const typingDiv = addTypingMessage();
            
            const formData = new FormData();
            formData.append('message', message);
            if (file) {
                formData.append('file', file);
            }
            
            fetch('/api/chat', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                removeTypingMessage();
                
                if (data.reply) {
                    const messageParts = processMessageWithCode(data.reply);
                    messageParts.forEach(part => {
                        if (part.type === 'text') {
                            addMessage('bot', part.content);
                        } else if (part.type === 'code') {
                            addMessage('bot', part.content, true, part.language);
                        }
                    });
                } else if (data.error) {
                    addMessage('bot', 'حدث خطأ: ' + data.error);
                }
            })
            .catch(error => {
                removeTypingMessage();
                addMessage('bot', 'حدث خطأ في الاتصال بالخادم');
                console.error('Error:', error);
            });
        }
        
        chatForm.addEventListener('submit', function(e) {
            e.preventDefault();
            const message = userInput.value.trim();
            if (message) {
                addMessage('user', message);
                userInput.value = '';
                sendMessageToServer(message);
            }
        });
        
        fileUpload.addEventListener('change', function(e) {
            if (this.files && this.files[0]) {
                const file = this.files[0];
                const fileName = file.name;
                
                addMessage('user', `تم تحميل الملف: ${fileName}`);
                sendMessageToServer(`تحليل الملف: ${fileName}`, file);
                
                // إعادة تعيين حقل التحميل
                this.value = '';
            }
        });
        
        newChatBtn.addEventListener('click', function() {
            if (confirm('هل تريد بدء محادثة جديدة؟ سيتم حذف محادثتك الحالية.')) {
                fetch('/api/new_chat', { method: 'POST' })
                    .then(() => {
                        chatMessages.innerHTML = '';
                        addMessage('bot', 'مرحباً! كيف يمكنني مساعدتك اليوم؟');
                    })
                    .catch(error => {
                        console.error('Error starting new chat:', error);
                    });
            }
        });
        
        // تمييز الكود عند التحميل
        document.querySelectorAll('pre code').forEach((block) => {
            hljs.highlightElement(block);
        });
    });
</script>
{% endblock %}
    ''',
    
    'features.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5">
    <div class="row">
        <div class="col-lg-8 mx-auto text-center">
            <h1 class="display-4 mb-4">ميزات OTH AI</h1>
            <p class="lead mb-5">اكتشف القوة الكاملة لمنصتنا الذكية</p>
        </div>
    </div>
    
    <div class="row g-4">
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-chat-square-text"></i></div>
                    <h3>محادثة ذكية</h3>
                    <p>تفاعل طبيعي مع نموذج Gemini 1.5 Flash الذي يفهم السياق ويقدم إجابات دقيقة.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-file-earmark-code"></i></div>
                    <h3>تحليل الكود</h3>
                    <p>اكتشف الأخطاء البرمجية، احصل على تفسيرات، وحسّن كودك بلغات متعددة.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-file-earmark-text"></i></div>
                    <h3>معالجة الملفات</h3>
                    <p>تحليل PDF، Word، Excel واستخراج المعلومات المهمة منها.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-image"></i></div>
                    <h3>تحليل الصور</h3>
                    <p>فهم المحتوى المرئي، قراءة النصوص في الصور، وتحليل المشاهد.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-translate"></i></div>
                    <h3>ترجمة متقدمة</h3>
                    <p>ترجمة النصوص بين اللغات مع الحفاظ على السياق والمعنى.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <div class="feature-icon mb-3"><i class="bi bi-lightbulb"></i></div>
                    <h3>إبداع المحتوى</h3>
                    <p>إنشاء مقالات، قصص، شعر، ونصوص إبداعية أخرى بلمسة بشرية.</p>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    ''',
    
    'about.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5">
    <div class="row">
        <div class="col-lg-8 mx-auto text-center">
            <h1 class="display-4 mb-4">حول OTH AI</h1>
            <p class="lead mb-5">منصة الذكاء الاصطناعي المتكاملة للجميع</p>
        </div>
    </div>
    
    <div class="row g-4">
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <h3 class="mb-3">من نحن</h3>
                    <p>OTH AI هي منصة ذكاء اصطناعي متقدمة تعتمد على نموذج Gemini 1.5 Flash من جوجل، مصممة لتقديم تجربة محادثة ذكية وسريعة للمستخدمين.</p>
                    <p>نهدف إلى جعل الذكاء الاصطناعي متاحاً وسهل الاستخدام للجميع، من المطورين إلى المستخدمين العاديين.</p>
                </div>
            </div>
        </div>
        <div class="col-md-6">
            <div class="card h-100">
                <div class="card-body">
                    <h3 class="mb-3">التقنية</h3>
                    <p>نستخدم أحدث نماذج الذكاء الاصطناعي من جوجل مع تحسينات خاصة لزيادة الدقة والسرعة.</p>
                    <p>المنصة مبنية بتقنيات حديثة تضمن الأمان والخصوصية وسهولة الاستخدام.</p>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    ''',
    
    '404.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5 text-center">
    <h1 class="display-1 text-danger">404</h1>
    <h2 class="mb-4">الصفحة غير موجودة</h2>
    <p class="lead">عذراً، الصفحة التي تبحث عنها غير موجودة.</p>
    <a href="/" class="btn btn-primary">العودة للصفحة الرئيسية</a>
</div>
{% endblock %}
    ''',
    
    'profile.html': '''
{% extends "base.html" %}

{% block content %}
<div class="container py-5">
    <div class="row">
        <div class="col-lg-8 mx-auto">
            <div class="card shadow">
                <div class="card-header bg-primary text-white">
                    <h4 class="mb-0"><i class="bi bi-person-circle me-2"></i> الملف الشخصي</h4>
                </div>
                <div class="card-body">
                    <div class="row">
                        <div class="col-md-4 text-center">
                            <div class="mb-3">
                                <i class="bi bi-person-bounding-box" style="font-size: 5rem;"></i>
                            </div>
                            <h5>{{ session['username'] }}</h5>
                            <p class="text-muted">عضو منذ {{ user_data.created_at | datetimeformat }}</p>
                        </div>
                        <div class="col-md-8">
                            <form>
                                <div class="mb-3">
                                    <label class="form-label">اسم المستخدم</label>
                                    <input type="text" class="form-control" value="{{ session['username'] }}" readonly>
                                </div>
                                <div class="mb-3">
                                    <label class="form-label">البريد الإلكتروني</label>
                                    <input type="email" class="form-control" value="{{ user_data.email or 'غير محدد' }}" readonly>
                                </div>
                                <div class="mb-3">
                                    <label class="form-label">تاريخ التسجيل</label>
                                    <input type="text" class="form-control" value="{{ user_data.created_at | datetimeformat }}" readonly>
                                </div>
                                <div class="mb-3">
                                    <label class="form-label">عدد المحادثات</label>
                                    <input type="text" class="form-control" value="{{ user_data.conversation_count or '0' }}" readonly>
                                </div>
                            </form>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
</div>
{% endblock %}
    '''
}

# وظائف إدارة البيانات
def load_users():
    global users
    try:
        if os.path.exists(DATA_FILE):
            with open(DATA_FILE, 'r') as f:
                users = json.load(f)
        else:
            users = {}
    except Exception as e:
        logger.error(f"Error loading users: {str(e)}")
        users = {}

def save_users():
    try:
        with open(DATA_FILE, 'w') as f:
            json.dump(users, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving users: {str(e)}")

# تحميل المستخدمين عند البدء
load_users()

# ديكورات المسارات
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('يجب تسجيل الدخول أولاً', 'danger')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# وظائف مساعدة
def get_user_id(sender_id):
    return hashlib.md5(sender_id.encode()).hexdigest()

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'pdf', 'doc', 'docx', 'xls', 'xlsx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_file(file):
    try:
        content_type = mimetypes.guess_type(file.filename)[0]
        
        if content_type.startswith('image/'):
            # معالجة الصور
            img = Image.open(file.stream)
            text = pytesseract.image_to_string(img, lang='ara+eng')
            return text
        
        elif content_type == 'application/pdf':
            # معالجة PDF
            images = convert_from_bytes(file.read())
            text = ""
            for img in images:
                text += pytesseract.image_to_string(img, lang='ara+eng') + "\n"
            return text
        
        elif content_type in ['application/msword', 'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
            # معالجة ملفات Word (يتطلب python-docx)
            import docx
            doc = docx.Document(file)
            return "\n".join([para.text for para in doc.paragraphs])
        
        elif content_type in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            # معالجة ملفات Excel (يتطلب openpyxl)
            from openpyxl import load_workbook
            wb = load_workbook(filename=file.stream)
            text = ""
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell]) + "\n"
            return text
        
    except Exception as e:
        logger.error(f"Error extracting text from file: {str(e)}")
        return None

def analyze_content(content, context=None, is_code=False):
    try:
        if is_code:
            prompt = "حلل الكود التالي وقدم شرحاً مفصلاً:\n\n" + content
            if context:
                prompt = f"سياق المحادثة:\n{context}\n{prompt}"
        else:
            prompt = "حلل المحتوى التالي:\n\n" + content
            if context:
                prompt = f"سياق المحادثة:\n{context}\n{prompt}"
        
        response = model.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error(f"Error analyzing content: {str(e)}")
        return None

def detect_programming_language(code):
    try:
        # تحليل الكود لاكتشاف اللغة البرمجية
        prompt = f"حدد اللغة البرمجية للكود التالي:\n\n{code}\n\nأجب فقط باسم اللغة بدون أي شرح إضافي."
        response = model.generate_content(prompt)
        return response.text.strip()
    except Exception as e:
        logger.error(f"Error detecting programming language: {str(e)}")
        return "unknown"

def cleanup_old_conversations():
    current_time = time.time()
    with data_lock:
        for user_id in list(conversations.keys()):
            if current_time - conversations[user_id]["last_active"] > CONVERSATION_TIMEOUT:
                del conversations[user_id]
                logger.info(f"تم حذف محادثة المستخدم {user_id} لانتهاء المهلة")

# تعديل دالة render_template لاستخدام القوالب المضمنة
def render_template(template_name, **context):
    if template_name in TEMPLATES:
        # إضافة السمة المظلمة إذا لم تكن محددة
        if 'theme' not in context:
            context['theme'] = session.get('theme', 'light')
        return render_template_string(TEMPLATES[template_name], **context)
    raise Exception(f"Template {template_name} not found")

# مسارات الويب
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        with data_lock:
            user = users.get(username)
            if user and check_password_hash(user['password'], password):
                session['user_id'] = user['id']
                session['username'] = username
                session['session_id'] = str(uuid.uuid4())
                session.permanent = True
                
                # إنشاء محادثة جديدة للمستخدم إذا لم تكن موجودة
                if user['id'] not in conversations:
                    conversations[user['id']] = {
                        "history": ["بدأ المستخدم محادثة جديدة"],
                        "last_active": time.time()
                    }
                
                flash('تم تسجيل الدخول بنجاح!', 'success')
                next_page = request.args.get('next')
                return redirect(next_page or url_for('chat'))
            else:
                flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'danger')
    
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if len(username) < 4 or len(password) < 6:
            flash('اسم المستخدم يجب أن يكون 4 أحرف على الأقل وكلمة المرور 6 أحرف', 'danger')
            return redirect(url_for('register'))
        
        with data_lock:
            if username in users:
                flash('اسم المستخدم موجود بالفعل', 'danger')
            else:
                user_id = str(uuid.uuid4())
                users[username] = {
                    'id': user_id,
                    'username': username,
                    'password': generate_password_hash(password),
                    'created_at': datetime.now().isoformat(),
                    'conversation_count': 0
                }
                save_users()
                
                # إنشاء محادثة جديدة للمستخدم
                conversations[user_id] = {
                    "history": ["بدأ المستخدم محادثة جديدة"],
                    "last_active": time.time()
                }
                
                flash('تم إنشاء الحساب بنجاح! يمكنك تسجيل الدخول الآن', 'success')
                return redirect(url_for('login'))
    
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    user_id = session.get('user_id')
    with data_lock:
        if user_id in conversations:
            del conversations[user_id]
    
    session.clear()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('home'))

@app.route('/chat')
@login_required
def chat():
    return render_template('chat.html')

@app.route('/api/chat', methods=['POST'])
@login_required
def api_chat():
    if 'user_id' not in session:
        return jsonify({"error": "غير مصرح به"}), 401
    
    try:
        user_id = session['user_id']
        user_message = request.form.get('message', '').strip()
        file = request.files.get('file')
        
        if not user_message and not file:
            return jsonify({"error": "الرجاء إدخال رسالة أو تحميل ملف"}), 400
        
        with data_lock:
            if user_id not in conversations:
                conversations[user_id] = {
                    "history": ["بدأ المستخدم محادثة جديدة"],
                    "last_active": time.time()
                }
            
            # تحديث وقت النشاط
            conversations[user_id]["last_active"] = time.time()
            
            # معالجة الملف إذا تم تحميله
            file_analysis = None
            if file and allowed_file(file.filename):
                file_text = extract_text_from_file(file)
                if file_text:
                    file_analysis = analyze_content(file_text, "\n".join(conversations[user_id]["history"][-5:]))
                    conversations[user_id]["history"].append(f"ملف: {file.filename}")
                    if file_analysis:
                        conversations[user_id]["history"].append(f"تحليل الملف: {file_analysis[:200]}...")
            
            # إضافة رسالة المستخدم
            if user_message:
                conversations[user_id]["history"].append(f"المستخدم: {user_message}")
            
            # الحصول على سياق المحادثة
            context = "\n".join(conversations[user_id]["history"][-5:])
            
            # توليد الرد
            if file_analysis:
                prompt = f"تحليل الملف:\n{file_analysis}\n\nالسؤال: {user_message}" if user_message else f"تحليل الملف:\n{file_analysis}"
            else:
                prompt = user_message
            
            # اكتشاف إذا كان السؤال يتعلق بالكود البرمجي
            is_code_related = any(keyword in user_message.lower() for keyword in ['كود', 'برمجة', 'برمج', 'code', 'programming'])
            
            if is_code_related and not file:
                response = model.generate_content([
                    "أنت مساعد برمجي خبير. قم بتحليل الكود التالي وتقديم شرح مفصل:",
                    prompt
                ])
            else:
                response = model.generate_content(prompt)
            
            reply = response.text
            
            # تحسين الرد إذا كان يحتوي على كود
            if '```' in reply:
                language = detect_programming_language(reply)
                reply = reply.replace('```', f'```{language}')
            
            # إضافة رد البوت
            conversations[user_id]["history"].append(f"البوت: {reply}")
            
            return jsonify({"reply": reply}), 200
            
    except Exception as e:
        logger.error(f"API Error: {str(e)}")
        return jsonify({"error": "حدث خطأ أثناء معالجة طلبك"}), 500

@app.route('/api/conversation')
@login_required
def api_conversation():
    if 'user_id' not in session:
        return jsonify({"error": "غير مصرح به"}), 401
    
    user_id = session['user_id']
    with data_lock:
        if user_id in conversations:
            messages = []
            for msg in conversations[user_id]["history"]:
                if msg.startswith("المستخدم:"):
                    messages.append({"sender": "user", "content": msg[9:].strip()})
                elif msg.startswith("البوت:"):
                    messages.append({"sender": "bot", "content": msg[6:].strip()})
            return jsonify({"messages": messages})
        return jsonify({"messages": []})

@app.route('/api/new_chat', methods=['POST'])
@login_required
def api_new_chat():
    if 'user_id' not in session:
        return jsonify({"error": "غير مصرح به"}), 401
    
    user_id = session['user_id']
    with data_lock:
        conversations[user_id] = {
            "history": ["بدأ المستخدم محادثة جديدة"],
            "last_active": time.time()
        }
        users[session['username']]['conversation_count'] += 1
        save_users()
    
    return jsonify({"status": "success"})

@app.route('/toggle_theme')
def toggle_theme():
    current_theme = session.get('theme', 'light')
    new_theme = 'dark' if current_theme == 'light' else 'light'
    session['theme'] = new_theme
    return redirect(request.referrer or url_for('home'))

@app.route('/features')
def features():
    return render_template('features.html')

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/profile')
@login_required
def profile():
    user_data = users.get(session['username'], {})
    return render_template('profile.html', user_data=user_data)

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

# تشغيل التنظيف الدوري كل ساعة
def periodic_cleanup():
    while True:
        time.sleep(3600)  # كل ساعة
        cleanup_old_conversations()
        save_users()  # حفظ بيانات المستخدمين بانتظام

# بدء التنظيف الدوري في خيط منفصل
cleanup_thread = Thread(target=periodic_cleanup)
cleanup_thread.daemon = True
cleanup_thread.start()

if __name__ == '__main__':
    app.run(threaded=True)
